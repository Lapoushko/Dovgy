import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


class Report:
    # Класс для создания Excel таблицы

    def __init__(self, profes, years, avgSalary, avgSalaryProf, countVacsYear,
                 countVacsYearProf, nameFile):
        self.years = years
        self.avgSalary = avgSalary
        self.avgSalaryProf = avgSalaryProf
        self.countVacsYear = countVacsYear
        self.countVacsYearProf = countVacsYearProf
        self.profession = profes
        self.nameFile = nameFile

    def generateExcel(self):
        dataFrame = [[self.years[i], self.avgSalary[i], self.avgSalaryProf[i],
               self.countVacsYear[i], self.countVacsYearProf[i]] for i in range(len(self.years))]
        dataFrame.insert(0, ["Год", "Средняя зарплата", f"Средняя зарплата - {self.profession}", "Количество вакансий", f"Количество вакансий - {self.profession}"])
        dataFrame = pd.DataFrame(dataFrame, columns=None)
        with pd.ExcelWriter(self.nameFile) as writer:
            dataFrame.to_excel(writer, sheet_name='Статистика по годам', index=False, header=False)
        wb = openpyxl.load_workbook(self.nameFile)
        workTable1 = wb["Статистика по годам"]
        thin = Side(border_style="thin")
        self.addNewBorder(workTable1, thin, len(self.years) + 2, ["A", "B", "C", "D", "E"])
        self.createMaxWidth(workTable1)
        wb.save(self.nameFile)

    def addNewBorder(self, worksheet, side, count_columns, rows):
        for i in range(1, count_columns):
            for row in rows:
                if i == 1:
                    worksheet[row + str(i)].alignment = Alignment(horizontal='left')
                    worksheet[row + str(i)].font = Font(bold=True)
                if worksheet[row + str(i)].internal_value is not None:
                    worksheet[row + str(i)].border = Border(top=side, bottom=side, left=side, right=side)

    def createMaxWidth(self, table):
        dims = {}
        for row in table.rows:
            for cell in row:
                if cell.value is not None:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                else:
                    dims[cell.column] = len(str(cell.value))
        for col, value in dims.items():
            table.column_dimensions[get_column_letter(col)].width = value + 2
