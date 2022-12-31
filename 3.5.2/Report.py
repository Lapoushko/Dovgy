import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


class Report:
    # Класс для создания Excel таблицы

    def createMaxWidth(self, table):
        dims = {}
        for r in table.rows:
            for mesh in r:
                if mesh.value is not None:
                    dims[mesh.column] = max((dims.get(mesh.column, 0), len(str(mesh.value))))
                else:
                    dims[mesh.column] = len(str(mesh.value))
        for col, value in dims.items():
            table.column_dimensions[get_column_letter(col)].width = value + 2
    def __init__(self, profession, years, salaryAvg, salaryProfAvg, countVacsYear,
                 countVacsYearProf, nameFile):
        self.years = years
        self.salaryAvg = salaryAvg
        self.salaryProfAvg = salaryProfAvg
        self.countVacsYear = countVacsYear
        self.countVacsYearProf = countVacsYearProf
        self.profession = profession
        self.nameFile = nameFile

    def generateNewExcel(self):
        dataFrame = [[self.years[i], self.salaryAvg[i], self.salaryProfAvg[i],
               self.countVacsYear[i], self.countVacsYearProf[i]] for i in range(len(self.years))]
        dataFrame.insert(0, ["Год", "Средняя зарплата", f"Средняя зарплата - {self.profession}", "Количество вакансий", f"Количество вакансий - {self.profession}"])
        dataFrame = pd.DataFrame(dataFrame, columns=None)
        with pd.ExcelWriter(self.nameFile) as writer:
            dataFrame.to_excel(writer, sheet_name='Статистика по годам', index=False, header=False)
        wb = openpyxl.load_workbook(self.nameFile)
        workTable1 = wb["Статистика по годам"]
        thin = Side(border_style="thin")
        self.addNewBorder(workTable1, thin, len(self.years) + 2, ["A", "B", "C", "D", "E"])
        self.createMaxWidth(workTable1)
        wb.save(self.nameFile)

    def addNewBorder(self, table, side, countCol, rows):
        for i in range(1, countCol):
            for r in rows:
                if i == 1:
                    table[r + str(i)].alignment = Alignment(horizontal='left')
                    table[r + str(i)].font = Font(bold=True)
                if table[r + str(i)].internal_value is not None:
                    table[r + str(i)].border = Border(top=side, bottom=side, left=side, right=side)
