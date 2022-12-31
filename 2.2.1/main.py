import matplotlib.pyplot as plt
import numpy as np
import csv
from operator import itemgetter
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00



def TakeItems(d):
    dNew = {}
    i = 0
    for key in d:
        dNew[key] = round(d[key], 4)
        i += 1
        if i == 10:
            break
    return dNew

class Report:
    def __init__(self, ds):
        self.profession = ds.profession
        self.years_list_headers = (
            "Год", "Средняя зарплата", f"Средняя зарплата - {profession}", "Количество вакансий",
            f"Количество вакансий - {profession}")
        self.years_list_columns = [[y for y in ds.salary_by_years],
                                   [val for val in ds.salary_by_years.values()],
                                   [val for val in ds.salary_by_years_for_profession.values()],
                                   [val for val in ds.vacancies_count_by_years.values()],
                                   [val for val in ds.vacancies_count_by_years_for_profession.values()]]

        self.cities_list_headers = ("Город", "Уровень зарплат", "", "Город", "Доля вакансий")
        self.cities_list_columns = [[city for city in ds.salary_by_cities],
                                    [val for val in ds.salary_by_cities.values()],
                                    ["" for i in range(len(ds.salary_by_cities))],
                                    [city for city in ds.vacancies_share_by_cities],
                                    [val for val in ds.vacancies_share_by_cities.values()]]

        self.years_list_widths = [2 + len(hd) for hd in self.years_list_headers]
        for i in range(len(self.years_list_columns)):
            for cell in self.years_list_columns[i]:
                self.years_list_widths[i] = max(len(str(cell)) + 2, self.years_list_widths[i])

        self.cities_list_widths = [ 2 + len(hd) for hd in self.cities_list_headers]
        for i in range(len(self.cities_list_columns)):
            for cell in self.cities_list_columns[i]:
                self.cities_list_widths[i] = max(len(str(cell)) + 2, self.cities_list_widths[i])

    def SeetBorder(self, ws, wd, hg):
        rangeCell = f'A1:{get_column_letter(wd)}{hg}'
        t = Side(border_style="thin", color="000000")
        for row in ws[rangeCell]:
            for cell in row:
                cell.border = Border(top=t, left=t, right=t, bottom=t)

    def ClearCol(self, ws, col):
        em = Side(border_style=None)
        for cell in ws[col]:
            cell.border = Border(top=em, bottom=em)

    def GetVertChart(self, title, param1, nameParam1, param2,
                     nameParam2, labels, fg, num):
        x = np.arange(len(labels))
        wd = 0.35
        plt.rcParams['font.size'] = '8'
        a_x = fg.add_subplot(num)
        a_x.bar(x - wd / 2, param1, wd, label=nameParam1)
        a_x.bar(x + wd / 2, param2, wd, label=nameParam2)
        a_x.set_xticks(x, labels, rotation="vertical")
        a_x.grid(axis='y')
        a_x.set_title(title)
        a_x.legend()

    def GetHorChart(self, title, param, labels, fg):
        plt.rcParams['font.size'] = '8'
        a_x = fg.add_subplot(223)
        labels = [city.replace(' ', '\n').replace('-', '-\n') for city in labels]
        y = np.arange(len(labels))
        a_x.barh(y, param)
        a_x.set_yticks(y, labels=labels, fontsize=6)
        a_x.grid(axis='x')
        a_x.invert_yaxis()
        a_x.set_title(title)

    def GetPieChart(self, title, param, labels, fg):
        plt.rcParams['font.size'] = '6'
        labels.insert(0, "Другие")
        param.insert(0, 1 - sum(param))
        a_x = fg.add_subplot(224)
        a_x.pie(param, labels=labels)
        a_x.axis('equal')
        a_x.set_title(title)
        fg.tight_layout()
        plt.savefig('graph.png')

    def generate_excel(self):
        wb = openpyxl.Workbook()
        listY = wb.active
        listY.title = "Статистика по годам"
        listC = wb.create_sheet("Статистика по городам")
        listY.append(self.years_list_headers)
        for cell in listY['1']:
            cell.font = Font(bold=True)
        for i in range(len(self.years_list_columns[0])):
            listY.append([column[i] for column in self.years_list_columns])
        listC.append(self.cities_list_headers)
        for cell in listC['1']:
            cell.font = Font(bold=True)
        for i in range(len(self.cities_list_columns[0])):
            listC.append([column[i] for column in self.cities_list_columns])
        for cell in listC['E']:
            cell.number_format = FORMAT_PERCENTAGE_00
        for i in range(1, 6):
            listY.column_dimensions[get_column_letter(i)].width = self.years_list_widths[i - 1]
            listC.column_dimensions[get_column_letter(i)].width = self.cities_list_widths[i - 1]
        self.SeetBorder(listY, len(self.years_list_headers), len(self.years_list_columns[0]) + 1)
        self.SeetBorder(listC, len(self.cities_list_headers), len(self.cities_list_columns[0]) + 1)
        Report.ClearCol(listC, 'C')
        wb.save('report.xlsx')

    def generate_image(self):
        fg = plt.figure()
        self.GetVertChart("Уровень зарплат по годам", self.years_list_columns[1], "средняя з/п",
                          self.years_list_columns[2], f"з/п {self.profession}", self.years_list_columns[0], fg,
                          221)
        self.GetVertChart("Количество вакансий по годам", self.years_list_columns[3], "Количество вакансий",
                          self.years_list_columns[4], f"Количество вакансий {self.profession}",
                          self.years_list_columns[0], fg, 222)
        self.GetHorChart("Уровень зарплат по городам", self.cities_list_columns[1],
                         self.cities_list_columns[0], fg)
        self.GetPieChart("Доля вакансий по городам", self.cities_list_columns[4], self.cities_list_columns[3], fg)
        fg.tight_layout()
        plt.savefig('graph.png')



class DataSet:
    def __init__(self, file_name, profession):
        self.file_name = file_name
        self.profession = profession
        headlines, vacancies = self.CSVReader()
        dictionaries = self.CSVFilter(vacancies, headlines)
        self.vacancies_objects = [Vacancy(dictionary) for dictionary in dictionaries]
        self.vacancies_count_by_years = self.GetVacCountY()
        self.vacancies_count_by_years_for_profession = self.GetVacCountYearsProf()
        self.salary_by_years = self.GetSalaryY()
        self.salary_by_years_for_profession = self.GetSalaryYProf()
        self.vacancies_count_by_cities = self.GetVacCities()
        self.vacancies_share_by_cities = self.GetShareCities()
        self.salary_by_cities = self.GetSalaryCities()

    def CSVReader(self):
        hdList = []
        vacList = []
        leng = 0
        isFirst = True
        countRows = 0
        with open(self.file_name, encoding="utf-8-sig") as File:
            rd = csv.reader(File)
            for row in rd:
                countRows += 1
                if isFirst:
                    hdList = row
                    leng = len(row)
                    isFirst = False
                else:
                    isBreak = False
                    if leng != len(row):
                        isBreak = True
                    for w in row:
                        if w == "":
                            isBreak = True
                    if isBreak:
                        continue
                    vacList.append(row)
        if countRows == 0:
            print("Пустой файл")
            exit()
        if countRows == 1:
            print("Нет данных")
            exit()
        return hdList, vacList

    def CSVFilter(self, rd, listNames):
        listDict = []
        for vac in rd:
            d = {}
            for i in range(len(listNames)):
                d[listNames[i]] = vac[i]
            listDict.append(d)
        return listDict

    def GetVacCountY(self):
        d = {}
        for vac in self.vacancies_objects:
            if vac.published_at in d:
                d[vac.published_at] += 1
            else:
                d[vac.published_at] = 1
        d = dict(sorted(d.items(), key=itemgetter(0)))
        return d

    def GetVacCountYearsProf(self):
        d = {}
        for vac in self.vacancies_objects:
            if self.profession not in vac.name:
                continue
            if vac.published_at in d:
                d[vac.published_at] += 1
            else:
                d[vac.published_at] = 1
        d = dict(sorted(d.items(), key=itemgetter(0)))
        if len(d) == 0:
            d = {2022: 0}
        return d

    def GetSalaryY(self):
        d = {}
        for vac in self.vacancies_objects:
            if vac.published_at in d:
                d[vac.published_at] += vac.salary
            else:
                d[vac.published_at] = vac.salary
        for key in d:
            d[key] = int(d[key] / self.vacancies_count_by_years[key])
        d = dict(sorted(d.items(), key=itemgetter(0)))
        return d

    def GetSalaryYProf(self):
        d = {}
        for vac in self.vacancies_objects:
            if self.profession not in vac.name:
                continue
            if vac.published_at in d:
                d[vac.published_at] += vac.salary
            else:
                d[vac.published_at] = vac.salary
        for key in d:
            d[key] = int(d[key] / self.vacancies_count_by_years_for_profession[key])
        d = dict(sorted(d.items(), key=itemgetter(0)))
        if len(d) == 0:
            d = {2022: 0}
        return d

    def GetVacCities(self):
        d = {}
        for vac in self.vacancies_objects:
            if vac.area_name in d:
                d[vac.area_name] += 1
            else:
                d[vac.area_name] = 1
        return d

    def GetShareCities(self):
        d = {}
        for key in self.vacancies_count_by_cities:
            if self.vacancies_count_by_cities[key] / len(self.vacancies_objects) >= 0.01:
                d[key] = self.vacancies_count_by_cities[key] / len(self.vacancies_objects)
        d = dict(sorted(d.items(), key=itemgetter(1), reverse=True))
        dNew = TakeItems(d)
        return dNew

    def GetSalaryCities(self):
        d = {}
        for vac in self.vacancies_objects:
            if self.vacancies_count_by_cities[vac.area_name] / len(self.vacancies_objects) < 0.01:
                continue
            if vac.area_name in d:
                d[vac.area_name] += vac.salary
            else:
                d[vac.area_name] = vac.salary
        for key in d:
            d[key] = int(d[key] / self.vacancies_count_by_cities[key])
        d = dict(sorted(d.items(), key=itemgetter(1), reverse=True))
        dNew = TakeItems(d)
        return dNew

    def PrintInfo(self):
        print("Динамика уровня зарплат по годам: " + str(self.salary_by_years))
        print("Динамика количества вакансий по годам: " + str(self.vacancies_count_by_years))
        print("Динамика уровня зарплат по годам для выбранной профессии: " + str(
            self.salary_by_years_for_profession))
        print("Динамика количества вакансий по годам для выбранной профессии: " + str(
            self.vacancies_count_by_years_for_profession))
        print("Уровень зарплат по городам (в порядке убывания): " + str(self.salary_by_cities))
        print("Доля вакансий по городам (в порядке убывания): " + str(self.vacancies_share_by_cities))


class Vacancy:
    def __init__(self, d):
        self.name = d["name"]
        self.salary = (float(d["salary_from"]) + float(d["salary_to"])) / 2 * cur[
            d["salary_currency"]]
        self.area_name = d["area_name"]
        self.published_at = int(d["published_at"][:4])



cur = {"AZN": 35.68,
                   "BYR": 23.91,
                   "EUR": 59.90,
                   "GEL": 21.74,
                   "KGS": 0.76,
                   "KZT": 0.13,
                   "RUR": 1,
                   "UAH": 1.64,
                   "USD": 60.66,
                   "UZS": 0.0055, }

nameFile = input("Введите название файла: ")
profession = input("Введите название профессии: ")
ds = DataSet(nameFile, profession)
ds.PrintInfo()
Report(ds).generate_image()
